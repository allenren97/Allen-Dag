"""Read the business intake spreadsheet into a list of normalized row dicts."""
from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook


REQUIRED_COLUMNS = (
    "business_line",
    "cadence",
    "engine",
    "connection_id_import",
    "connection_id_export",
    "database",
    "schema",
    "table",
)


def _normalize(value: Any) -> Any:
    """Trim strings and coerce blank / 'None' literals to ``None``."""
    if isinstance(value, str):
        v = value.strip()
        if v == "" or v.lower() == "none":
            return None
        return v
    return value


def read_intake_rows(intake_path: Path) -> list[dict[str, Any]]:
    """
    Return the data rows of the first worksheet as a list of dicts keyed by
    header. Empty rows and rows missing ``business_line`` / ``cadence`` are
    skipped. Raises ``ValueError`` if a required column is missing.
    """
    wb = load_workbook(intake_path, data_only=True, read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    try:
        header_raw = next(rows_iter)
    except StopIteration:
        return []
    header = [str(c).strip() if c is not None else "" for c in header_raw]

    missing = [c for c in REQUIRED_COLUMNS if c not in header]
    if missing:
        raise ValueError(
            f"Intake header is missing required column(s): {missing}"
        )

    rows: list[dict[str, Any]] = []
    for raw in rows_iter:
        if raw is None or all(cell is None for cell in raw):
            continue
        row = {key: _normalize(value) for key, value in zip(header, raw)}
        if not row.get("business_line") or not row.get("cadence"):
            continue
        rows.append(row)
    return rows
